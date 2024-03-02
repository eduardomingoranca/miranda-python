# openpyxl - manipulando as planilhas do Workbook
# Com essa biblioteca sera possivel ler e escrever dados em celulas
# especificas, formatar celulas, inserir graficos,
# criar formulas, adicionar imagens e outros elementos graficos as suas
# planilhas. Ela eh util para automatizar tarefas envolvendo planilhas do
# Excel, como a criacao de relatorios e analise de dados e/ou facilitando a
# manipulacao de grandes quantidades de informacoes.
# Instalacao necessaria: pip install openpyxl
# Documentacao: https://openpyxl.readthedocs.io/en/stable/
from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

workbook = Workbook()
# worksheet: Worksheet = workbook.active

# Nome para a planilha
sheet_name = 'Minha planilha'
# Criamos a planilha
workbook.create_sheet(sheet_name, 0)
# Selecionou a planilha
worksheet: Worksheet = workbook[sheet_name]

# Remover uma planilha
workbook.remove(workbook['Sheet'])

# Criando os cabecalhos
worksheet.cell(1, 1, 'Nome')
worksheet.cell(1, 2, 'Idade')
worksheet.cell(1, 3, 'Nota')

students = [
    # nome      idade nota
    ['Joao',    14,   5.5],
    ['Maria',   13,   9.7],
    ['Luiz',    15,   8.8],
    ['Alberto', 16,   10],
]

# for i, student_row in enumerate(students, start=2):
#     for j, student_column in enumerate(student_row, start=1):
#         worksheet.cell(i, j, student_column)

for student in students:
    worksheet.append(student)

workbook.save(WORKBOOK_PATH)