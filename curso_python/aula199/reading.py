# openpyxl - ler e alterar dados de uma planilha
# Com essa biblioteca sera possivel ler e escrever dados em celulas
# especificas, formatar celulas, inserir graficos,
# criar formulas, adicionar imagens e outros elementos graficos as suas
# planilhas. Ela eh util para automatizar tarefas envolvendo planilhas do
# Excel, como a criacao de relatorios e analise de dados e/ou facilitando a
# manipulacao de grandes quantidades de informacoes.
# Instalacao necessaria: pip install openpyxl
# Documentacao: https://openpyxl.readthedocs.io/en/stable/
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

# Carregando um arquivo do excel
workbook: Workbook = load_workbook(WORKBOOK_PATH)

# Nome para a planilha
sheet_name = 'Minha planilha'

# Selecionou a planilha
worksheet: Worksheet = workbook[sheet_name]

row: tuple[Cell]
for row in worksheet.iter_rows(min_row=2):
    for cell in row:
        print(cell.value, end='\t')

        if cell.value == 'Mary':
            worksheet.cell(cell.row, 2, 23)
    print()

# worksheet['B3'].value = 14

workbook.save(WORKBOOK_PATH)